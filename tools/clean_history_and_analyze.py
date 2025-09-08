"""Robust cleaner for History.xlsx and portfolio analyzer.

- Finds the header row by searching for common column names (date, ticker, symbol, quantity, price)
- Builds a pandas DataFrame from rows below header
- Cleans and normalizes columns
- Writes history_cleaned.csv and history_analysis.txt at repo root
"""
import os
from openpyxl import load_workbook
import pandas as pd
import yfinance as yf
from datetime import datetime
import requests
import urllib.parse
import json

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
HISTORY_XLSX = os.path.join(ROOT, 'History.xlsx')
OUT_CSV = os.path.join(ROOT, 'history_cleaned.csv')
OUT_ANALYSIS = os.path.join(ROOT, 'history_analysis.txt')
LOG = os.path.join('tools', 'history_clean_and_analyze.log')

common_cols = ['date', 'trade_date', 'ticker', 'symbol', 'quantity', 'qty', 'shares', 'price', 'fill_price']

with open(LOG, 'w', encoding='utf-8') as log:
    log.write(f'Loading {HISTORY_XLSX}\n')
    if not os.path.exists(HISTORY_XLSX):
        """Robust cleaner for History.xlsx and portfolio analyzer.

        - Finds the header row by searching for common column names (date, ticker, symbol, quantity, price)
        - Builds a pandas DataFrame from rows below header
        - Cleans and normalizes columns
        - Writes history_cleaned.csv and history_analysis.txt at repo root
        """
        import os
        import argparse
        from openpyxl import load_workbook
        import pandas as pd
        import yfinance as yf
        from datetime import datetime
        import requests
        import urllib.parse
        import json

        ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        HISTORY_XLSX = os.path.join(ROOT, 'History.xlsx')
        OUT_CSV = os.path.join(ROOT, 'history_cleaned.csv')
        OUT_ANALYSIS = os.path.join(ROOT, 'history_analysis.txt')
        LOG = os.path.join('tools', 'history_clean_and_analyze.log')

        common_cols = ['date', 'trade_date', 'ticker', 'symbol', 'quantity', 'qty', 'shares', 'price', 'fill_price']

        parser = argparse.ArgumentParser()
        parser.add_argument('--force-yfinance', action='store_true', help='Skip Polygon and use yfinance for all price fetches')
        args = parser.parse_args()
        FORCE_YFINANCE = bool(args.force_yfinance)

        with open(LOG, 'w', encoding='utf-8') as log:
            log.write(f'Loading {HISTORY_XLSX}\n')
            if not os.path.exists(HISTORY_XLSX):
                log.write('History.xlsx not found. Exiting.\n')
                raise SystemExit('History.xlsx not found')

            wb = load_workbook(HISTORY_XLSX, read_only=True, data_only=True)
            sheet_name = wb.sheetnames[0]
            log.write(f'Opening sheet: {sheet_name}\n')
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            log.write(f'Total rows read: {len(rows)}\n')

            # find header row: look for a row that contains any of 'ticker','symbol','security id','date'
            header_idx = None
            header_row = None
            for i, r in enumerate(rows[:min(200, len(rows))]):
                if not r:
                    continue
                row_lower = [str(x).strip().lower() if x is not None else '' for x in r]
                if any(k in row_lower for k in ['ticker','symbol','security id','security_id','date','trade_date']):
                    header_idx = i
                    header_row = r
                    break
            log.write(f'Header idx: {header_idx}\n')
            if header_idx is None:
                # fallback: assume row 0 is header
                header_idx = 0
                header_row = rows[0] if rows else []
                log.write('Header not found automatically; using first row as header.\n')

            # build dataframe from rows after header
            data_rows = rows[header_idx+1:]
            headers = [str(c).strip() if c is not None else f'col_{i}' for i, c in enumerate(header_row)]
            df = pd.DataFrame(data_rows, columns=headers)
            # drop columns that are fully empty (often the left-most empty column)
            df = df.dropna(axis=1, how='all')
            log.write(f'DF initial shape: {df.shape}\n')

            # Drop fully-empty rows
            df = df.dropna(how='all')
            log.write(f'After dropping empty rows: {df.shape}\n')

            # Normalize column names
            cols = {c: c.strip().lower().replace(' ', '_') for c in df.columns}
            df.rename(columns=cols, inplace=True)

            # Map common columns
            col_map = {}
            if 'date' in df.columns:
                col_map['date'] = 'date'
            elif 'trade_date' in df.columns:
                col_map['date'] = 'trade_date'

            for candidate in ['ticker','symbol','security_id','security id','security id']:
                if candidate in df.columns:
                    col_map['ticker'] = candidate
                    break

            for candidate in ['quantity','qty','shares']:
                if candidate in df.columns:
                    col_map['quantity'] = candidate
                    break

            for candidate in ['price','fill_price']:
                if candidate in df.columns:
                    col_map['price'] = candidate
                    break

            # Rename to standard names
            rename_map = {v: k for k, v in col_map.items()}
            log.write(f'Rename map: {rename_map}\n')
            df = df.rename(columns=rename_map)

            # Keep useful columns first
            keep = [c for c in ['date','ticker','quantity','price'] if c in df.columns]
            other_cols = [c for c in df.columns if c not in keep]
            df = df[keep + other_cols]
            log.write(f'Columns after reorder: {df.columns.tolist()}\n')

            # Drop rows where ticker is missing
            if 'ticker' in df.columns:
                df = df[~df['ticker'].isna()]
                df['ticker'] = df['ticker'].astype(str).str.strip().str.upper().str.replace('.', '-', regex=False)

            # Parse dates
            if 'date' in df.columns:
                df['date'] = pd.to_datetime(df['date'], errors='coerce')

            # Drop rows without date
            if 'date' in df.columns:
                before = len(df)
                df = df[~df['date'].isna()]
                log.write(f'Dropped {before - len(df)} rows without valid date\n')

            # Try to coerce numeric columns
            if 'quantity' in df.columns:
                df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce')
            if 'price' in df.columns:
                df['price'] = pd.to_numeric(df['price'], errors='coerce')

            # Drop rows with no ticker or date
            df = df.dropna(subset=[c for c in ['ticker','date'] if c in df.columns])

            df = df.drop_duplicates()

            # Save cleaned CSV
            df.to_csv(OUT_CSV, index=False)
            log.write(f'Wrote cleaned CSV to {OUT_CSV} with {len(df)} rows\n')

            # Analysis
            analysis_lines = []
            if 'quantity' in df.columns and 'price' in df.columns:
                df['trade_value'] = df['quantity'] * df['price']
                agg = df.groupby('ticker').agg(
                    total_quantity=('quantity','sum'),
                    total_cost=('trade_value','sum'),
                    trades=('ticker','count')
                ).reset_index()
                # avoid division by zero
                agg['avg_cost'] = agg.apply(lambda r: (r['total_cost']/r['total_quantity']) if r['total_quantity'] else 0, axis=1)

                analysis_lines.append('Portfolio positions (by ticker):')
                analysis_lines.append(agg.to_string(index=False))

                tickers = [t for t in agg['ticker'].tolist() if isinstance(t, str) and t.strip()]
                if FORCE_YFINANCE:
                    log.write(f'Fetching live prices for {len(tickers)} tickers (yfinance only)\n')
                else:
                    log.write(f'Fetching live prices for {len(tickers)} tickers (Polygon -> yfinance fallback)\n')
                prices = {}
                # Load Polygon API key from environment or .env
                API_KEY = os.getenv('POLYGON_API_KEY')
                if not API_KEY:
                    env_path = os.path.join(ROOT, '.env')
                    if os.path.exists(env_path):
                        try:
                            with open(env_path, 'r', encoding='utf-8') as ef:
                                for L in ef:
                                    if L.strip().startswith('POLYGON_API_KEY='):
                                        API_KEY = L.strip().split('=', 1)[1].strip()
                                        break
                        except Exception:
                            log.write('Failed to read .env for POLYGON_API_KEY\n')

                # Allow overriding the Polygon files host (S3 endpoint) via env
                FILES_HOST = os.getenv('POLYGON_FILES_HOST', 'https://files.polygon.io')
                for t in tickers:
                    p = None
                    # Try the Polygon 'files' S3 host first if API key available and not forcing yfinance
                    if API_KEY and not FORCE_YFINANCE:
                        try:
                            url_files = f"{FILES_HOST}/v2/last/trade/{urllib.parse.quote(t)}?apiKey={API_KEY}"
                            resp = requests.get(url_files, timeout=10)
                            if resp.status_code == 200:
                                j = resp.json()
                                r = j.get('results') or j.get('result')
                                if isinstance(r, dict):
                                    p = r.get('p') if r.get('p') is not None else r.get('price')
                            else:
                                log.write(f'Files host returned {resp.status_code} for {t} ({FILES_HOST})\n')
                                # If files host didn't work, try the normal API host
                                url_api = f"https://api.polygon.io/v2/last/trade/{urllib.parse.quote(t)}?apiKey={API_KEY}"
                                resp2 = requests.get(url_api, timeout=10)
                                if resp2.status_code == 200:
                                    j = resp2.json()
                                    r = j.get('results') or j.get('result')
                                    if isinstance(r, dict):
                                        p = r.get('p') if r.get('p') is not None else r.get('price')
                                elif resp2.status_code in (401, 403):
                                    log.write(f'Polygon auth error {resp2.status_code} for {t}\n')
                                    API_KEY = None
                                else:
                                    log.write(f'API host returned {resp2.status_code} for {t}\n')
                        except Exception as e:
                            log.write(f'Polygon fetch error for {t}: {e}\n')

                    # Fallback to yfinance per-ticker if Polygon did not return a price or when forced
                    if p is None:
                        try:
                            hist = yf.Ticker(t).history(period='1d')
                            if not hist.empty and 'Close' in hist.columns:
                                p = hist['Close'].dropna().iloc[-1]
                        except Exception as e:
                            log.write(f'yfinance fetch failed for {t}: {e}\n')

                    prices[t] = p

                agg['market_price'] = agg['ticker'].map(prices)
                agg['market_value'] = agg['market_price'] * agg['total_quantity']
                agg['unrealized_pnl'] = agg['market_value'] - agg['total_cost']
                agg['unrealized_pct'] = (agg['unrealized_pnl'] / agg['total_cost']) * 100

                analysis_lines.append('\nMarket snapshot:')
                analysis_lines.append(agg[['ticker','total_quantity','avg_cost','market_price','market_value','unrealized_pnl','unrealized_pct']].to_string(index=False))

                total_cost = agg['total_cost'].sum()
                total_market = agg['market_value'].sum()
                total_pnl = total_market - total_cost
                total_pct = (total_pnl / total_cost) * 100 if total_cost else 0

                analysis_lines.append('\nPortfolio summary:')
                analysis_lines.append(f'Total cost basis: ${total_cost:,.2f}')
                analysis_lines.append(f'Total market value: ${total_market:,.2f}')
                analysis_lines.append(f'Total unrealized P&L: ${total_pnl:,.2f} ({total_pct:.2f}%)')
            else:
                analysis_lines.append('Insufficient columns for position analysis (need quantity and price).')

            with open(OUT_ANALYSIS, 'w', encoding='utf-8') as f:
                f.write('\n'.join(analysis_lines))
            log.write(f'Wrote analysis to {OUT_ANALYSIS}\n')

        print('Done - outputs:')
        print(OUT_CSV)
        print(OUT_ANALYSIS)
        print('Log:', LOG)
